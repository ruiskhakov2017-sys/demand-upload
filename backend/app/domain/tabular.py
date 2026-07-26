from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

MAX_IMPORT_ROWS = 5000


def parse_tabular(filename: str, content: bytes) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Поддерживаются только CSV и XLSX")
    if not rows:
        raise ValueError("В файле нет строк с данными")
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"В одном импорте допускается не более {MAX_IMPORT_ROWS} строк")
    return [_normalize_row(row) for row in rows]


def _parse_csv(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1251")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("В CSV отсутствует строка заголовков")
    return [dict(row) for row in reader if any(str(value or "").strip() for value in row.values())]


def _parse_xlsx(content: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(iterator)]
    except StopIteration as exc:
        raise ValueError("В XLSX нет строк") from exc
    if not any(headers):
        raise ValueError("В XLSX отсутствует строка заголовков")
    rows: list[dict] = []
    for values in iterator:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(str(value or "").strip() for value in row.values()):
            rows.append(row)
    return rows


ALIASES = {
    "account": "customer_id",
    "account_id": "customer_id",
    "customer": "customer_id",
    "customer id": "customer_id",
    "аккаунт": "customer_id",
    "аккаунт id": "customer_id",
    "campaign": "campaign_name",
    "campaign name": "campaign_name",
    "кампания": "campaign_name",
    "название кампании": "campaign_name",
    "url": "final_url",
    "final url": "final_url",
    "ссылка": "final_url",
    "budget": "daily_budget",
    "daily budget": "daily_budget",
    "бюджет": "daily_budget",
    "headline": "headlines",
    "заголовки": "headlines",
    "description": "descriptions",
    "описания": "descriptions",
    "youtube": "youtube_video_id",
    "youtube id": "youtube_video_id",
}


def _normalize_row(row: dict) -> dict:
    normalized: dict = {}
    for raw_key, raw_value in row.items():
        key = " ".join(str(raw_key or "").strip().lower().replace("_", " ").split())
        key = ALIASES.get(key, key.replace(" ", "_"))
        value = raw_value.strip() if isinstance(raw_value, str) else raw_value
        if key in {"headlines", "descriptions", "location_ids", "language_ids", "media_ids"}:
            if isinstance(value, str):
                value = [item.strip() for item in value.replace("\n", "|").split("|") if item.strip()]
        normalized[key] = value
    return normalized
